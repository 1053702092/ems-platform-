#!/usr/bin/env python3
"""Update the priority Excel with new research findings."""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy
import sys, io

SRC = "docs/岗位/00_EMS_BMS_岗位_优先投递精简版.xlsx"
DST = "docs/岗位/00_EMS_BMS_岗位_优先投递精简版.xlsx"

wb = openpyxl.load_workbook(SRC)

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

def find_row(ws, keyword):
    """Find first row where col G (company name) contains keyword."""
    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[6] and keyword in str(row[6]):
            return ri
    return None

def update_cell(ws, row, col, value, fill=None):
    """Update cell value preserving existing formatting."""
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if fill:
        cell.fill = fill

# =====================================================
# 1. 央企国企优先 sheet - update existing + add new
# =====================================================
ws_中央 = wb["央企国企优先"]

# 1a. 中国电气装备集团储能科技 (Row 11) - confirmed 校招, upgrade score
r = find_row(ws_中央, "电气装备")
if r:
    update_cell(ws_中央, r, 3, 92)   # 综合分: 83→92
    update_cell(ws_中央, r, 4, "高")  # 契合度
    update_cell(ws_中央, r, 5, "较高") # Offer概率
    update_cell(ws_中央, r, 9, "上海（本部）/许昌/天津/济南")
    update_cell(ws_中央, r, 10, "campus.51job.com/cnkj2026\n刘老师 021-66698846")
    update_cell(ws_中央, r, 11, "✅ 确定有2026校招！招~30人。EMS软件工程师，七险二金，上海落户。储能央企新设，高度对口。")
    update_cell(ws_中央, r, 12, "校招已开，抓紧投递。特别优秀可免笔试。")
    update_cell(ws_中央, r, 13, "✅ 立即投递")
    print(f"Updated: 中国电气装备集团储能科技 (Row {r}) -> score 92")

# 1b. 广东新型储能国家研究院 (Row 32) - mostly PhD, adjust down
r = find_row(ws_中央, "新型储能国家研究院")
if r:
    update_cell(ws_中央, r, 3, 78)   # 综合分: 83→78
    update_cell(ws_中央, r, 11, "⚠️ 以博士/博士后为主，硕士可投部分副研究员岗(19-30k·15薪)。门槛高但薪资诱人。")
    update_cell(ws_中央, r, 10, "zhaopin@naesic.com")
    update_cell(ws_中央, r, 13, "⚠️ 冲刺项，硕可试")
    print(f"Updated: 广东新型储能国家研究院 (Row {r}) -> score 78")

# 1c. NEW: 广州储能集团 (insert at top)
new_rows_中央 = [
    # (priority, batch, score, fit, offer_prob, category, company, direction, location, channel, why, risk, next, source)
    (None, "P0 立即优先", 93, "高", "较高", "央企国企优先",
     "广州储能集团有限公司",
     "产品研发助理工程师（储能EMS方向）",
     "广州",
     "手机端 http://bjxapp.cn/t/NjMwODE3OA/",
     "✅ 确定有2026校招！2023年成立的新国企，注册资本20亿，广州储能'链主'企业。产品研发岗直接对口EMS项目。新公司校招第一年门槛最低。",
     "新公司，薪资待遇待核实",
     "✅ 立即投递",
     "来源=web-research-2026-07-29；笔记=广州储能集团有明确校招"),

    (None, "P0 立即优先", 88, "高", "较高", "央企国企优先",
     "中国资源循环集团有限公司",
     "技术专责/管培生/工艺工程师（新能源/电池方向）",
     "天津（总部）/深圳/南京/杭州",
     "https://zgzh.iguopin.com",
     "✅ 确定有2026校招！2025年新央企，国务院国资委直接监管。化学/材料专业在这里是优势不是短板！深圳有子公司。资源循环是政策强推赛道。",
     "新央企，业务方向还在搭建中",
     "✅ 立即投递",
     "来源=web-research-2026-07-29；笔记=中国资环2026校招确认"),
]

# Insert after row 1 (after header), shifting others down
insert_row = 2
for i, row_data in enumerate(new_rows_中央):
    ws_中央.insert_rows(insert_row + i)
    for ci, val in enumerate(row_data):
        if val is not None:
            ws_中央.cell(row=insert_row + i, column=ci + 1, value=val)
    # Set priority number
    ws_中央.cell(row=insert_row + i, column=1, value=f"{insert_row + i}")
    # Highlight green
    for ci in range(1, 15):
        ws_中央.cell(row=insert_row + i, column=ci).fill = GREEN

print(f"Inserted: 广州储能集团 + 中国资源循环集团 into 央企国企优先")

# Renumber priorities in 央企国企优先
for ri in range(2, ws_中央.max_row + 1):
    ws_中央.cell(row=ri, column=1, value=ri - 1)

# =====================================================
# 2. 大厂外企冲刺 sheet - update CATL-related
# =====================================================
ws_大厂 = wb["大厂外企冲刺"]

r = find_row(ws_大厂, "CATL智能科技")
if r:
    update_cell(ws_大厂, r, 3, 72)   # 69→72
    update_cell(ws_大厂, r, 10, "Talent.CATL.com\ncatl-campus@catl.com")
    update_cell(ws_大厂, r, 11, "CATL 2026校招进行中！BMS软件与算法/储能系统开发岗，20-50w/年。专业门槛仍在，但今年明确写了化学/材料可投。")
    update_cell(ws_大厂, r, 13, "可投，别抱太大期望")
    print(f"Updated: CATL (Row {r}) -> score 72")

r = find_row(ws_大厂, "新能时代储能")
if r:
    update_cell(ws_大厂, r, 3, 68)   # 65→68
    update_cell(ws_大厂, r, 11, "CATL系独立储能子公司，门槛比CATL总部低。但校招信息不明确，需关注。")
    print(f"Updated: 新能时代储能 (Row {r}) -> score 68")

# =====================================================
# 3. 南方城市优先 sheet - add new entries
# =====================================================
ws_南方 = wb["南方城市优先"]

# Find last row with data
last_row = ws_南方.max_row + 1

new_南方 = [
    (None, "P0 立即优先", 93, "高", "较高", "央企国企优先",
     "广州储能集团有限公司",
     "产品研发助理工程师",
     "广州",
     "http://bjxapp.cn/t/NjMwODE3OA/",
     "✅ 校招确认！广州储能国企，EMS方向对口，新公司门槛低", "", "✅ 立即投递", ""),
    (None, "P0 立即优先", 88, "高", "较高", "央企国企优先",
     "中国资源循环集团有限公司",
     "技术专责/管培生",
     "深圳/南京/杭州",
     "https://zgzh.iguopin.com",
     "✅ 校招确认！化学/材料对口，深圳有岗，新央企", "", "✅ 立即投递", ""),
    (None, "P0 立即优先", 78, "中", "中", "央企国企优先",
     "广东新型储能国家研究院",
     "副研究员（储能控制/仿真）",
     "广州",
     "zhaopin@naesic.com",
     "⚠️ 硕可投部分岗位，19-30k·15薪，研究院环境", "", "⚠️ 冲刺项", ""),
]

for i, row_data in enumerate(new_南方):
    r = last_row + i
    for ci, val in enumerate(row_data):
        if val is not None:
            ws_南方.cell(row=r, column=ci + 1, value=val)
    ws_南方.cell(row=r, column=1, value=r - 1)
    for ci in range(1, 15):
        ws_南方.cell(row=r, column=ci).fill = GREEN if i < 2 else YELLOW

print(f"Added {len(new_南方)} entries to 南方城市优先")

# =====================================================
# 4. 优先投递总表 - Add new entries
# =====================================================
ws_总 = wb["优先投递总表"]
last_row_总 = ws_总.max_row + 1

new_总 = [
    (None, "P0 立即优先", 93, "高", "较高", "央企国企优先",
     "广州储能集团有限公司",
     "产品研发助理工程师（储能EMS方向）",
     "广州",
     "http://bjxapp.cn/t/NjMwODE3OA/",
     "✅ 校招确认！广州储能国企，EMS直接对口", "", "✅ 立即投递", ""),
    (None, "P0 立即优先", 92, "高", "较高", "央企国企优先",
     "中国电气装备集团储能科技",
     "EMS软件工程师（30万/年）",
     "上海",
     "campus.51job.com/cnkj2026",
     "✅ 校招确认！七险二金，高度对口", "", "✅ 立即投递", ""),
    (None, "P0 立即优先", 88, "高", "较高", "央企国企优先",
     "中国资源循环集团有限公司",
     "技术专责/管培生",
     "深圳/南京/杭州",
     "https://zgzh.iguopin.com",
     "✅ 校招确认！化学/材料对口", "", "✅ 立即投递", ""),
]

for i, row_data in enumerate(new_总):
    r = last_row_总 + i
    for ci, val in enumerate(row_data):
        if ci < len(row_data) and val is not None:
            ws_总.cell(row=r, column=ci + 1, value=val)
    ws_总.cell(row=r, column=1, value=r - 1)
    for ci in range(1, 15):
        ws_总.cell(row=r, column=ci).fill = GREEN

print(f"Added {len(new_总)} entries to 优先投递总表")

# =====================================================
# Save
# =====================================================
wb.save(DST)
print(f"\nOK -> Updated {DST}")
